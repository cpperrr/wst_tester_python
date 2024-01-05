import os
import pathlib
import shutil

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook


# Class to handle writing of SP logfiles, status files etc.
# noinspection PyBroadException
class SpFileWriter:
	def __init__(self):
		self.cell_width_margin = 1.0

	@staticmethod
	def create_backup_file(file_path):
		file_path = pathlib.Path(file_path)
		original_file_path = pathlib.Path(file_path)
		file_extension = file_path.suffix
		i = 0
		while file_path.is_file():
			file_path = original_file_path.with_suffix("%s.bak_%s" % (file_extension, i))
			i += 1
		shutil.copy(original_file_path, file_path)
		if file_path.is_file():
			return True
		else:
			return False

	@staticmethod
	def add_file_extension_if_missing(file_path, extension):
		file_path = pathlib.Path(file_path)
		return file_path.with_suffix(extension)

	# noinspection PyBroadException
	def write_log_to_excel(self, parsed_log, excel_file_path, overwrite=False):

		excel_file_path = self.add_file_extension_if_missing(excel_file_path, extension='.xlsx')

		if pathlib.Path(excel_file_path).is_file():
			if overwrite:
				if self.create_backup_file(excel_file_path):
					os.remove(excel_file_path)  # we only remove file if backup was created.
			else:
				raise Exception("File already exists. use overwrite=True to overwrite.")

		wb = Workbook()
		ws = wb.active
		ws.title = "Log"

		# Ready data
		log_headings = parsed_log['headings']
		log_records = parsed_log['records']

		# write headings and log
		ws = wb.active
		ws.append(log_headings)
		for record in log_records:
			ws.append(record)

		# Style the sheet
		# Style headings
		ft = Font(bold=True)
		for col in ws["A1:%s1" % get_column_letter(ws.max_column)]:
			for cell in col:
				cell.font = ft

		# Style dates
		# date format (2022-08-25 08:41:54)
		date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM:SS')
		for date_cell in ws['B']:
			if date_cell.row == 1:
				continue
			date_cell.style = date_style

		# Adjust width of columns
		# Iterate over all columns and adjust their widths
		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = (max_length + 2) * self.cell_width_margin
			ws.column_dimensions[column_letter].width = adjusted_width

		# Apply filter
		full_cell_range = "A1:" + get_column_letter(ws.max_column) + str(ws.max_row)
		ws.auto_filter.ref = full_cell_range

		# save file
		wb.save(excel_file_path)

	def write_log_stats_to_excel(self, parsed_log_statistics, excel_file_path, append=True):
		excel_file_path = self.add_file_extension_if_missing(excel_file_path, extension='.xlsx')

		if not pathlib.Path(excel_file_path).is_file():
			if append:
				raise Exception("Excel file does not exist. Cannot append to it.")
			else:
				wb = Workbook()
				ws = wb.active
				ws.title = "Log Statistics"
		else:
			wb = load_workbook(excel_file_path)
			ws = wb.create_sheet(title="Log Statistics")

		# write statistics
		ws.append(["Name", "Value"])
		for record in parsed_log_statistics.items():
			ws.append(record)

		# Adjust width of columns
		# Iterate over all columns and adjust their widths

		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = (max_length + 2) * self.cell_width_margin
			ws.column_dimensions[column_letter].width = adjusted_width

		wb.save(excel_file_path)

	def write_general_info_to_excel(self, general_battery_info, excel_file_path, append=True):
		excel_file_path = self.add_file_extension_if_missing(excel_file_path, extension='.xlsx')

		if not pathlib.Path(excel_file_path).is_file():
			if append:
				raise Exception("Excel file does not exist. Cannot append to it.")
			else:
				wb = Workbook()
				ws = wb.active
				ws.title = "Log Statistics"
		else:
			wb = load_workbook(excel_file_path)
			ws = wb.create_sheet(title="Battery Information")

		# write statistics
		ws.append(["Name", "Value"])
		for record in general_battery_info.items():
			ws.append(record)

		# Adjust width of columns
		# Iterate over all columns and adjust their widths

		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = (max_length + 2) * self.cell_width_margin
			ws.column_dimensions[column_letter].width = adjusted_width

		wb.save(excel_file_path)

	def write_realtime_status_to_excel(self, realtime_status, excel_file_path, append=True):
		excel_file_path = self.add_file_extension_if_missing(excel_file_path, extension='.xlsx')

		if not pathlib.Path(excel_file_path).is_file():
			if append:
				raise Exception("Excel file does not exist. Cannot append to it.")
			else:
				wb = Workbook()
				ws = wb.active
				ws.title = "Real-time status"
		else:
			wb = load_workbook(excel_file_path)
			ws = wb.create_sheet(title="Realtime status")

		data_to_write = []
		for key, value in realtime_status['parsed_voltage_status'].items():
			if isinstance(value, dict):
				for subkey, sub_value in value.items():
					if type(sub_value) in (tuple, list):
						if not sub_value:
							sub_value = ""
						else:
							sub_value = str(sub_value)
					subkey = str(subkey).capitalize().replace("_", " ")
					data_to_write.append([subkey, sub_value])
			else:
				if type(value) in (tuple, list):
					if not value:
						value = ""
					else:
						value = str(value)
				subkey = str(key).capitalize().replace("_", " ")
				data_to_write.append([subkey, value])

		current_row = 0
		for key, value in realtime_status['parsed_current_status'].items():
			if isinstance(value, dict):
				for subkey, sub_value in value.items():
					if type(sub_value) in (tuple, list):
						if not sub_value:
							sub_value = ""
						else:
							sub_value = str(sub_value)
					subkey = str(subkey).capitalize().replace("_", " ")
					if current_row >= len(data_to_write):
						data_to_write.append(["", ""])
					data_to_write[current_row] += ["", subkey, sub_value]
					current_row += 1
			else:
				if type(value) in (tuple, list):
					if not value:
						value = ""
					else:
						value = str(value)
				subkey = str(key).capitalize().replace("_", " ")
				if current_row >= len(data_to_write):
					data_to_write.append(["", ""])
				data_to_write[current_row] += ["", subkey, value]
				current_row += 1

		current_row = 0
		for key, value in realtime_status['parsed_power_status'].items():
			if isinstance(value, dict):
				for subkey, sub_value in value.items():
					if type(sub_value) in (tuple, list):
						if not sub_value:
							sub_value = ""
						else:
							sub_value = str(sub_value)
					subkey = str(subkey).capitalize().replace("_", " ")
					if current_row >= len(data_to_write):
						data_to_write.append(["", "", "", ""])
					data_to_write[current_row] += ["", subkey, sub_value]
					current_row += 1
			else:
				if type(value) in (tuple, list):
					if not value:
						value = ""
					else:
						value = str(value)
				subkey = str(key).capitalize().replace("_", " ")
				if current_row >= len(data_to_write):
					data_to_write.append(["", "", "", ""])
				data_to_write[current_row] += ["", subkey, value]
				current_row += 1

		for row in data_to_write:
			ws.append(row)

		# Adjust width of columns
		# Iterate over all columns and adjust their widths

		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = (max_length + 2) * self.cell_width_margin
			ws.column_dimensions[column_letter].width = adjusted_width

		wb.save(excel_file_path)

	def write_parameters_to_excel(self, parsed_parameters, excel_file_path, append=True):
		excel_file_path = self.add_file_extension_if_missing(excel_file_path, extension='.xlsx')

		if not pathlib.Path(excel_file_path).is_file():
			if append:
				raise Exception("Excel file does not exist. Cannot append to it.")
			else:
				wb = Workbook()
				ws = wb.active
				ws.title = "Parameters"
		else:
			wb = load_workbook(excel_file_path)
			ws = wb.create_sheet(title="Parameters")

		data_to_write = []
		indent = 0
		for group, parameters_in_group in parsed_parameters.items():

			current_row = 0
			for param_name, param_value_dict in parameters_in_group.items():

				if current_row >= len(data_to_write):
					data_to_write.append(["", "", ""]*indent + ["%s(%s)" % (param_name, param_value_dict['unit']), param_value_dict['value'], ""])
				else:

					extra_indent = ["", "", ""] * (indent - int(len(data_to_write[current_row])/3))
					data_to_write[current_row] = data_to_write[current_row] + extra_indent + ["%s(%s)" % (param_name, param_value_dict['unit']), param_value_dict['value'], ""]
				current_row += 1
			indent += 1

		# Write the data
		for row in data_to_write:
			ws.append(row)

		# Adjust width of columns
		# Iterate over all columns and adjust their widths

		for column in ws.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = (max_length + 2) * self.cell_width_margin
			ws.column_dimensions[column_letter].width = adjusted_width

		wb.save(excel_file_path)
